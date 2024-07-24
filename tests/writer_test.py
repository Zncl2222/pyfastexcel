from __future__ import annotations

import pytest
from openpyxl.styles import Side
from openpyxl_style_writer import CustomStyle

from pyfastexcel import StreamWriter

font_params = {
    'size': 11,
    'bold': True,
    'italic': True,
    'color': 'FF000000',
    'vertAlign': 'baseline',
    'strike': True,
    'name': 'Calibri',
    'family': 1,
    'underline': 'doubleAccounting',
}

fill_params = {
    'fill_type': 'solid',
    'start_color': 'FFFFFFFF',
    'end_color': 'FF000000',
}

border_params = {
    'left': Side(style='thin', color='FF000000'),
    'right': Side(style='thick', color='FF000000'),
    'top': Side(style='dotted', color='FF000000'),
    'bottom': Side(style='dashDot', color='FF000000'),
    'diagonal': Side(style='hair', color='FF000000'),
    'diagonal_direction': 1,
    'outline': Side(style='medium', color='FF000000'),
    'vertical': Side(style='mediumDashed', color='FF000000'),
    'horizontal': Side(style='slantDashDot', color='FF000000'),
}

ali_params = {
    'horizontal': 'general',
    'vertical': 'bottom',
    'text_rotation': 12,
    'wrap_text': True,
    'shrink_to_fit': True,
    'indent': 1,
    'justifyLastLine': True,
    'readingOrder': 1,
    'relativeIndent': 1,
}


def prepare_example_data(rows: int = 1000, cols: int = 10) -> list[dict[str, str]]:
    headers = [f'Column_{i}' for i in range(cols)]
    data = [[i for i in range(cols)] for j in range(rows)]
    records = []
    for row in data:
        record = {}
        for header, value in zip(headers, row):
            record[header] = str(round(value * 100, 2))
        records.append(record)
    return records


class StyleCollections:
    black_fill_style = CustomStyle(
        font_name='Time News Roman',
        font_size='11',
        font_bold=True,
        font_color='F62B00',
        fill_color='000000',
    )
    green_fill_style = CustomStyle(
        font_size='29',
        font_bold=False,
        font_color='000000',
        fill_color='375623',
    )
    test_fill_style = CustomStyle(
        font_params={
            'size': 20,
            'bold': True,
            'italic': True,
            'color': '5e03fc',
        },
        fill_params={
            'patternType': 'solid',
            'fgColor': '375623',
        },
        border_params={
            'left': Side(style='thin', color='e12aeb'),
            'right': Side(style='thick', color='e12aeb'),
            'top': Side(style=None, color='e12aeb'),
            'bottom': Side(style='dashDot', color='e12aeb'),
        },
        ali_params={
            'wrapText': True,
            'shrinkToFit': True,
        },
        number_format='0.00%',
    )
    test_style = CustomStyle(
        font_params=font_params,
        fill_params=fill_params,
        border_params=border_params,
        ali_params=ali_params,
        number_format='0.00%',
        protect=True,
    )
    test_style.protection.hidden = True


class PyFastExcelStreamExample(StreamWriter, StyleCollections):
    def create_excel(self) -> bytes:
        self._set_header()
        self._create_style()
        self.set_file_props('Creator', 'Hello')
        self._create_single_header()
        self._create_body()
        return self.read_lib_and_create_excel()

    def _set_header(self):
        self.headers = list(self.data[0].keys())
        self.set_cell_height(self.sheet, 5, 12)
        self.set_cell_width(self.sheet, 'A', 12)
        self.set_cell_width(self.sheet, 3, 12)

    def _create_single_header(self):
        for h in self.headers:
            self.row_append(h, style='green_fill_style')
        self.create_row()

    def _create_body(self) -> None:
        for row in self.data:
            for h in self.headers:
                if h[-1] in ('1', '3', '5', '7', '9'):
                    self.row_append(row[h], style=self.black_fill_style)
                else:
                    self.row_append(row[h], style='test_fill_style')
            self.create_row()

        self.create_sheet('Sheet2')
        for row in self.data:
            for h in self.headers:
                if h[-1] in ('1', '3', '5', '7', '9'):
                    self.row_append(row[h], style=self.green_fill_style)
                else:
                    self.row_append(row[h], style='black_fill_style')
            self.create_row()

        self.workbook['Sheet1']['A4'] = 'Test with default style'
        self.workbook['Sheet1']['A3'] = ('Hello', 'test_style')

        # Test Local Style
        custom_style2 = CustomStyle(
            font_size='33',
            font_bold=True,
            font_color='000000',
            fill_color='4db3af',
        )
        self.row_append('Local Style', style=custom_style2)
        self.create_row()

        custom_style3 = CustomStyle(
            font_size='33',
            font_bold=True,
            font_color='000000',
            fill_color='4db3af',
        )
        self.row_append('Local Style', style=custom_style3)
        self.create_row()

        # Test non-numeric value for 'validate_and_format_value'
        self.row_append(['1', 2, 3])
        self.create_row()

        # AutoFilter
        self.auto_filter('Sheet1', 'A1:C1')


def test_set_data_with_index():
    excel_example = PyFastExcelStreamExample([[None] * 1000 for _ in range(1000)])
    excel_example.workbook['Sheet1']['A1'] = 'test'
    excel_example.workbook['Sheet1']['AZ4455'] = 'I am 1234!!!'

    with pytest.raises(TypeError):
        excel_example.workbook['Sheet1']['A1'] = ('test', [])


def test_get_data_with_index():
    excel_example = PyFastExcelStreamExample([[None] * 1000 for _ in range(1000)])
    excel_example.workbook['Sheet1']['A1'] = 'test'
    print(excel_example.workbook['Sheet1']['A1'])

    with pytest.raises(TypeError):
        excel_example.workbook['Sheet1']['A1'] = ('test', [])


def test_set_data_with_cell():
    excel_example = PyFastExcelStreamExample([[None] * 1000 for _ in range(1000)])
    ws = excel_example.workbook['Sheet1']
    ws.cell(row=1, column=1, value='test')
    ws.cell(row=12312, column=11221, value='I am 1234!!!')

    with pytest.raises(TypeError):
        ws.cell(row=1, column=1, value=('test', []))

    with pytest.raises(ValueError):
        ws.cell(row=199999999, column=1, value='test')

    with pytest.raises(ValueError):
        ws.cell(row=1, column=99999999, value='test')


def test_set_file_props():
    excel_example = PyFastExcelStreamExample([[None] * 1000 for _ in range(1000)])
    with pytest.raises(ValueError):
        excel_example.set_file_props('Test', 'Test')


@pytest.mark.parametrize(
    'sheet, expected_exception',
    [
        ('Sheet1', ValueError),  # Invalid case
        ('Sheet2', None),  # Valid case
        ('Sheet3', None),  # Valid case
    ],
)
def test_create_sheet(sheet, expected_exception):
    excel_example = PyFastExcelStreamExample([[None] * 1000 for _ in range(1000)])
    if expected_exception is None:
        excel_example.create_sheet(sheet)
    else:
        with pytest.raises(expected_exception):
            excel_example.create_sheet(sheet)


def test_remove_sheet():
    excel_example = PyFastExcelStreamExample([[None] * 1000 for _ in range(1000)])
    with pytest.raises(ValueError):
        excel_example.remove_sheet('Sheet1')
    excel_example.create_sheet('Sheet2')
    excel_example.remove_sheet('Sheet2')

    excel_example.create_sheet('Sheet3')
    with pytest.raises(IndexError):
        excel_example.remove_sheet('Sheet333')


@pytest.mark.parametrize(
    'sheet, column, width, expected_exception',
    [
        ('Sheet1', 16385, 12, ValueError),  # Invalid case
        ('qwe', '', '', KeyError),  # Invalid: Single cell is not a merge cell
    ],
)
def test_set_cell_width(sheet, column, width, expected_exception):
    excel_example = PyFastExcelStreamExample([])
    with pytest.raises(expected_exception):
        excel_example.set_cell_width(sheet, column, width)


@pytest.mark.parametrize(
    'sheet, row, height, expected_exception',
    [
        ('Sheet1', 1048577, 12, ValueError),  # Invalid case
        ('qwe', '', '', KeyError),  # Invalid: Single cell is not a merge cell
    ],
)
def test_set_cell_height(sheet, row, height, expected_exception):
    excel_example = PyFastExcelStreamExample([[None] * 1000 for _ in range(1000)])
    with pytest.raises(expected_exception):
        excel_example.set_cell_height(sheet, row, height)


@pytest.mark.parametrize(
    'sheet, top_left_cell, bottom_right_cell, expected_exception',
    [
        ('Sheet1', 'A1', 'C2', None),  # Valid case
        ('Sheet1', 'A1', 'A1', None),  # Invalid: Single cell is not a merge cell
        ('Sheet1', 'A1048577', 'C2', ValueError),  # Invalid: Row number exceeds limit
        ('Sheet1', 'A1', 'C1048577', ValueError),  # Invalid: Row number exceeds limit
        ('Sheet1', 'XFD1', 'XFD1048576', None),  # Valid: Maximum row and column numbers
        ('Sheet1', 'A1', 'XFE1048576', ValueError),  # Invalid: Column number exceeds limit
        ('Sheet1', 'A2', 'A1', ValueError),  # Invalid: Top number less than bottom number
        ('Sheet1', 'C1', 'A1', ValueError),  # Invalid: Top column less than bottom column
        ('Sheet1', 'A0', 'A1', ValueError),  # Invalid: Row number too small
        ('Sheet1', 'A0', 'C0', ValueError),  # Invalid: Row number too small
        ('abcd', '', '', KeyError),  # Invalid: Sheet name not found
    ],
)
def test_set_merge_cell(sheet, top_left_cell, bottom_right_cell, expected_exception):
    excel = PyFastExcelStreamExample([[None] * 1000 for _ in range(1000)])
    if expected_exception is not None:
        with pytest.raises(expected_exception):
            excel.merge_cell(sheet, top_left_cell, bottom_right_cell)
    else:
        # set_merge_cell will be remove in v1.0.0, use merge_cell instead
        excel.set_merge_cell(sheet, top_left_cell, bottom_right_cell)
        assert (top_left_cell, bottom_right_cell) in excel.workbook[sheet].merge_cells


@pytest.mark.parametrize(
    'sheet, cell_range, expected_exception',
    [
        ('Sheet1', 'A1:C2', None),  # Valid case
        ('Sheet1', 'A1:A1', None),  # Invalid: Single cell is not a merge cell
        ('Sheet1', 'A1:A1', ValueError),  # Invalid: Single cell is not a merge cell
    ],
)
def test_set_merge_cell_with_cell_range(sheet, cell_range, expected_exception):
    excel = PyFastExcelStreamExample([[None] * 1000 for _ in range(1000)])
    if expected_exception is not None:
        with pytest.raises(expected_exception):
            excel.merge_cell(sheet, cell_range)
    else:
        # set_merge_cell will be remove in v1.0.0, use merge_cell instead
        excel.set_merge_cell(sheet, cell_range=cell_range)
        top_left_cell = cell_range.split(':')[0]
        bottom_right_cell = cell_range.split(':')[1]
        assert (top_left_cell, bottom_right_cell) in excel.workbook[sheet].merge_cells


def test_pyfastexcel_stream_example():
    data = prepare_example_data(rows=3, cols=3)
    excel_example = PyFastExcelStreamExample(data)
    excel_example.create_sheet('Test')
    excel_example.remove_sheet('Test')
    excel_example.switch_sheet('Sheet1')
    excel_bytes = excel_example.create_excel()
    assert isinstance(excel_bytes, bytes)
    assert excel_example._dict_wb['Sheet2']['Data'][-1] == [("['1', 2, 3]", 'DEFAULT_STYLE')]
