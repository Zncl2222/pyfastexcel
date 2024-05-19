from __future__ import annotations

import pytest
from openpyxl.styles import Side
from openpyxl_style_writer import CustomStyle

from pyfastexcel import FastWriter, NormalWriter, Workbook

font_params = {
    'size': 11,
    'bold': True,
    'italic': True,
    'color': 'FF000000',
    'vertAlign': 'baseline',
    'strike': True,
    'name': 'Calibri',
    'family': 1,
    'underline': 'doubleAccounting',
}

fill_params = {
    'fill_type': 'solid',
    'start_color': 'FFFFFFFF',
    'end_color': 'FF000000',
}

border_params = {
    'left': Side(style='thin', color='FF000000'),
    'right': Side(style='thick', color='FF000000'),
    'top': Side(style='dotted', color='FF000000'),
    'bottom': Side(style='dashDot', color='FF000000'),
    'diagonal': Side(style='hair', color='FF000000'),
    'diagonal_direction': 1,
    'outline': Side(style='medium', color='FF000000'),
    'vertical': Side(style='mediumDashed', color='FF000000'),
    'horizontal': Side(style='slantDashDot', color='FF000000'),
}

ali_params = {
    'horizontal': 'general',
    'vertical': 'bottom',
    'text_rotation': 12,
    'wrap_text': True,
    'shrink_to_fit': True,
    'indent': 1,
    'justifyLastLine': True,
    'readingOrder': 1,
    'relativeIndent': 1,
}


def prepare_example_data(rows: int = 1000, cols: int = 10) -> list[dict[str, str]]:
    headers = [f'Column_{i}' for i in range(cols)]
    data = [[i for i in range(cols)] for j in range(rows)]
    records = []
    for row in data:
        record = {}
        for header, value in zip(headers, row):
            record[header] = str(round(value * 100, 2))
        records.append(record)
    return records


style_for_set_custom_style = CustomStyle(font_color='fcfcfc')


class StyleCollections:
    black_fill_style = CustomStyle(
        font_name='Time News Roman',
        font_size='11',
        font_bold=True,
        font_color='F62B00',
        fill_color='000000',
    )
    green_fill_style = CustomStyle(
        font_size='29',
        font_bold=False,
        font_color='000000',
        fill_color='375623',
    )
    test_fill_style = CustomStyle(
        font_params={
            'size': 20,
            'bold': True,
            'italic': True,
            'color': '5e03fc',
        },
        fill_params={
            'patternType': 'solid',
            'fgColor': '375623',
        },
        border_params={
            'left': Side(style='thin', color='e12aeb'),
            'right': Side(style='thick', color='e12aeb'),
            'top': Side(style=None, color='e12aeb'),
            'bottom': Side(style='dashDot', color='e12aeb'),
        },
        ali_params={
            'wrapText': True,
            'shrinkToFit': True,
        },
        number_format='0.00%',
    )
    test_style = CustomStyle(
        font_params=font_params,
        fill_params=fill_params,
        border_params=border_params,
        ali_params=ali_params,
        number_format='0.00%',
        protect=True,
    )
    test_style.protection.hidden = True


class PyExcelizeFastExample(FastWriter, StyleCollections):
    def create_excel(self) -> bytes:
        self._set_header()
        self._create_style()
        self._create_single_header()
        self._create_body()
        return self.read_lib_and_create_excel()

    def _set_header(self):
        self.headers = list(self.data[0].keys())
        self.set_cell_height(self.sheet, 5, 12)
        self.set_cell_width(self.sheet, 'A', 12)
        self.set_cell_width(self.sheet, 3, 12)

    def _create_single_header(self):
        for h in self.headers:
            self.row_append(h, style='green_fill_style')
        self.create_row()

    def _create_body(self) -> None:
        for row in self.data:
            for h in self.headers:
                if h[-1] in ('1', '3', '5', '7', '9'):
                    self.row_append(row[h], style='black_fill_style')
                else:
                    self.row_append(row[h], style='test_fill_style')
            self.create_row()

        self.create_sheet('Sheet2')
        self.switch_sheet('Sheet2')
        for row in self.data:
            for h in self.headers:
                if h[-1] in ('1', '3', '5', '7', '9'):
                    self.row_append(row[h], style='test_fill_style')
                else:
                    self.row_append(row[h], style=self.black_fill_style)
            self.create_row()
        self.workbook['Sheet1']['A4'] = 'Test with default style'
        self.workbook['Sheet1']['A3'] = ('Hello', 'test_style')

        # Test Local Style
        custom_style2 = CustomStyle(
            font_size='33',
            font_bold=True,
            font_color='000000',
            fill_color='4db3af',
        )
        self.row_append('Local Style', style=custom_style2)
        self.create_row()


class PyExcelizeNormalExample(NormalWriter, StyleCollections):
    def create_excel(self) -> bytes:
        self._set_header()
        self._create_style()
        self.set_file_props('Creator', 'Hello')
        self._create_single_header()
        self._create_body()
        return self.read_lib_and_create_excel()

    def _set_header(self):
        self.headers = list(self.data[0].keys())
        self.set_cell_height(self.sheet, 5, 12)
        self.set_cell_width(self.sheet, 'A', 12)
        self.set_cell_width(self.sheet, 3, 12)

    def _create_single_header(self):
        for h in self.headers:
            self.row_append(h, style='green_fill_style')
        self.create_row()

    def _create_body(self) -> None:
        for row in self.data:
            for h in self.headers:
                if h[-1] in ('1', '3', '5', '7', '9'):
                    self.row_append(row[h], style=self.black_fill_style)
                else:
                    self.row_append(row[h], style='test_fill_style')
            self.create_row()

        self.create_sheet('Sheet2')
        for row in self.data:
            for h in self.headers:
                if h[-1] in ('1', '3', '5', '7', '9'):
                    self.row_append(row[h], style=self.green_fill_style)
                else:
                    self.row_append(row[h], style='black_fill_style')
            self.create_row()

        custom_style2 = CustomStyle(
            font_size='33',
            font_bold=True,
            font_color='000000',
            fill_color='4db3af',
        )
        self.row_append('Local Style', style=custom_style2)
        self.create_row()

        # Test non-numeric value for 'validate_and_format_value'
        self.row_append(['1', 2, 3])
        self.create_row()


def test_pyexcelize_fast_example():
    from pyfastexcel.utils import set_custom_style

    set_custom_style('test', style_for_set_custom_style)
    data = prepare_example_data(rows=25, cols=9)
    excel_example = PyExcelizeFastExample(data)
    excel_bytes = excel_example.create_excel()
    assert isinstance(excel_bytes, bytes)


def test_set_data_with_index():
    excel_example = PyExcelizeFastExample([[None] * 1000 for _ in range(1000)])
    excel_example.workbook['Sheet1']['A1'] = 'test'
    excel_example.workbook['Sheet1']['AZ4455'] = 'I am 1234!!!'

    with pytest.raises(TypeError):
        excel_example.workbook['Sheet1']['A1'] = ('test', [])


def test_get_data_with_index():
    excel_example = PyExcelizeFastExample([[None] * 1000 for _ in range(1000)])
    excel_example.workbook['Sheet1']['A1'] = 'test'
    print(excel_example.workbook['Sheet1']['A1'])

    with pytest.raises(TypeError):
        excel_example.workbook['Sheet1']['A1'] = ('test', [])


def test_set_data_with_cell():
    excel_example = PyExcelizeFastExample([[None] * 1000 for _ in range(1000)])
    ws = excel_example.workbook['Sheet1']
    ws.cell(row=1, column=1, value='test')
    ws.cell(row=12312, column=11221, value='I am 1234!!!')

    with pytest.raises(TypeError):
        ws.cell(row=1, column=1, value=('test', []))

    with pytest.raises(ValueError):
        ws.cell(row=199999999, column=1, value='test')

    with pytest.raises(ValueError):
        ws.cell(row=1, column=99999999, value='test')


def test_set_data_faield_with_index():
    excel_example = PyExcelizeNormalExample([])
    with pytest.raises(IndexError):
        excel_example.workbook['Sheet1']['A1'] = 'qwe'
    with pytest.raises(IndexError):
        print(excel_example['Sheet1']['A1'])


def test_set_file_props():
    excel_example = PyExcelizeFastExample([[None] * 1000 for _ in range(1000)])
    with pytest.raises(ValueError):
        excel_example.set_file_props('Test', 'Test')


@pytest.mark.parametrize(
    'sheet, expected_exception',
    [
        ('Sheet1', ValueError),  # Invalid case
        ('Sheet2', None),  # Valid case
        ('Sheet3', None),  # Valid case
    ],
)
def test_create_sheet(sheet, expected_exception):
    excel_example = PyExcelizeFastExample([[None] * 1000 for _ in range(1000)])
    if expected_exception is None:
        excel_example.create_sheet(sheet)
    else:
        with pytest.raises(expected_exception):
            excel_example.create_sheet(sheet)


def test_remove_sheet():
    excel_example = PyExcelizeFastExample([[None] * 1000 for _ in range(1000)])
    with pytest.raises(ValueError):
        excel_example.remove_sheet('Sheet1')
    excel_example.create_sheet('Sheet2')
    excel_example.remove_sheet('Sheet2')

    excel_example.create_sheet('Sheet3')
    with pytest.raises(IndexError):
        excel_example.remove_sheet('Sheet333')


@pytest.mark.parametrize(
    'sheet, column, width, expected_exception',
    [
        ('Sheet1', 16385, 12, ValueError),  # Invalid case
        ('qwe', '', '', KeyError),  # Invalid: Single cell is not a merge cell
    ],
)
def test_set_cell_width(sheet, column, width, expected_exception):
    excel_example = PyExcelizeNormalExample([])
    with pytest.raises(expected_exception):
        excel_example.set_cell_width(sheet, column, width)


@pytest.mark.parametrize(
    'sheet, row, height, expected_exception',
    [
        ('Sheet1', 1048577, 12, ValueError),  # Invalid case
        ('qwe', '', '', KeyError),  # Invalid: Single cell is not a merge cell
    ],
)
def test_set_cell_height(sheet, row, height, expected_exception):
    excel_example = PyExcelizeFastExample([[None] * 1000 for _ in range(1000)])
    with pytest.raises(expected_exception):
        excel_example.set_cell_height(sheet, row, height)


@pytest.mark.parametrize(
    'sheet, top_left_cell, bottom_right_cell, expected_exception',
    [
        ('Sheet1', 'A1', 'C2', None),  # Valid case
        ('Sheet1', 'A1', 'A1', None),  # Invalid: Single cell is not a merge cell
        ('Sheet1', 'A1048577', 'C2', ValueError),  # Invalid: Row number exceeds limit
        ('Sheet1', 'A1', 'C1048577', ValueError),  # Invalid: Row number exceeds limit
        ('Sheet1', 'XFD1', 'XFD1048576', None),  # Valid: Maximum row and column numbers
        ('Sheet1', 'A1', 'XFE1048576', ValueError),  # Invalid: Column number exceeds limit
        ('Sheet1', 'A2', 'A1', ValueError),  # Invalid: Top number less than bottom number
        ('Sheet1', 'C1', 'A1', ValueError),  # Invalid: Top column less than bottom column
        ('Sheet1', 'A0', 'A1', ValueError),  # Invalid: Row number too small
        ('Sheet1', 'A0', 'C0', ValueError),  # Invalid: Row number too small
        ('abcd', '', '', KeyError),  # Invalid: Sheet name not found
    ],
)
def test_set_merge_cell(sheet, top_left_cell, bottom_right_cell, expected_exception):
    excel = PyExcelizeFastExample([[None] * 1000 for _ in range(1000)])
    if expected_exception is not None:
        with pytest.raises(expected_exception):
            excel.set_merge_cell(sheet, top_left_cell, bottom_right_cell)
    else:
        excel.set_merge_cell(sheet, top_left_cell, bottom_right_cell)
        assert (top_left_cell, bottom_right_cell) in excel.workbook[sheet].merge_cells


def test_pyexcelize_normal_example():
    data = prepare_example_data(rows=3, cols=3)
    excel_example = PyExcelizeNormalExample(data)
    excel_example.create_sheet('Test')
    excel_example.remove_sheet('Test')
    excel_example.switch_sheet('Sheet1')
    excel_bytes = excel_example.create_excel()
    assert isinstance(excel_bytes, bytes)
    assert excel_example._dict_wb['Sheet2']['Data'][-1] == [("['1', 2, 3]", 'DEFAULT_STYLE')]


@pytest.mark.parametrize(
    'range_slice, values, expected_output',
    [
        (
            slice('A1', 'G1'),
            [1, 2, 3, 9, 8, 45, 11],
            [
                (1, 'DEFAULT_STYLE'),
                (2, 'DEFAULT_STYLE'),
                (3, 'DEFAULT_STYLE'),
                (9, 'DEFAULT_STYLE'),
                (8, 'DEFAULT_STYLE'),
                (45, 'DEFAULT_STYLE'),
                (11, 'DEFAULT_STYLE'),
            ],
        ),
        (slice('A1', 'B1'), [1, 2], [(1, 'DEFAULT_STYLE'), (2, 'DEFAULT_STYLE')]),
    ],
)
def test_workbook(range_slice, values, expected_output):
    wb = Workbook()
    ws = wb['Sheet1']
    ws[range_slice] = values

    actual_output = [tuple([cell for cell in row]) for row in ws[range_slice]]
    assert actual_output == expected_output


@pytest.mark.parametrize(
    'input_data, expected_exception',
    [
        ([slice('A1', 'G3'), [1, 2, 3]], ValueError),  # Invalid slice assignment
        ((1.62, [1]), TypeError),  # Invalid row assignment
    ],
)
def test_invalid_assignment(input_data, expected_exception):
    wb = Workbook()
    ws = wb['Sheet1']

    with pytest.raises(expected_exception):
        ws[input_data[0]] = input_data[1]


@pytest.mark.parametrize(
    'row_slice, value_list, expected_output',
    [
        (
            slice('A1', 'E1'),
            [2, 6, 7, 8, 9],
            [
                (2, 'DEFAULT_STYLE'),
                (6, 'DEFAULT_STYLE'),
                (7, 'DEFAULT_STYLE'),
                (8, 'DEFAULT_STYLE'),
                (9, 'DEFAULT_STYLE'),
            ],
        ),
        (
            slice('B6', 'F6'),
            ['qwe', 6, 7, -8, 'hello'],
            [
                ('', 'DEFAULT_STYLE'),
                ('qwe', 'DEFAULT_STYLE'),
                (6, 'DEFAULT_STYLE'),
                (7, 'DEFAULT_STYLE'),
                (-8, 'DEFAULT_STYLE'),
                ('hello', 'DEFAULT_STYLE'),
            ],
        ),
        (
            slice('E100', 'I100'),
            [('qwe', 'bold_font_style'), 6, 7, -8, 'hello'],
            [
                ('', 'DEFAULT_STYLE'),
                ('', 'DEFAULT_STYLE'),
                ('', 'DEFAULT_STYLE'),
                ('', 'DEFAULT_STYLE'),
                ('qwe', 'bold_font_style'),
                (6, 'DEFAULT_STYLE'),
                (7, 'DEFAULT_STYLE'),
                (-8, 'DEFAULT_STYLE'),
                ('hello', 'DEFAULT_STYLE'),
            ],
        ),
        (slice('A1', 'G100'), [1, 2, 3], ValueError),
    ],
)
def test_workbook_slice(row_slice, value_list, expected_output):
    wb = Workbook()
    ws = wb['Sheet1']

    if not isinstance(expected_output, list):
        with pytest.raises(expected_output):
            ws[row_slice] = value_list
        with pytest.raises(expected_output):
            print(ws[row_slice])
    else:
        ws[row_slice] = value_list
        print(ws[row_slice])
        assert ws[row_slice] == expected_output


@pytest.mark.parametrize(
    'index, value_list, expected_output',
    [
        (
            0,
            [2, 6, 7, 8, 9],
            [
                (2, 'DEFAULT_STYLE'),
                (6, 'DEFAULT_STYLE'),
                (7, 'DEFAULT_STYLE'),
                (8, 'DEFAULT_STYLE'),
                (9, 'DEFAULT_STYLE'),
            ],
        ),
        (
            1,
            ['qwe', 6, 7, -8, 'hello'],
            [
                ('qwe', 'DEFAULT_STYLE'),
                (6, 'DEFAULT_STYLE'),
                (7, 'DEFAULT_STYLE'),
                (-8, 'DEFAULT_STYLE'),
                ('hello', 'DEFAULT_STYLE'),
            ],
        ),
        (
            36669,
            [('qwe', 'bold_font_style'), [6, 7, 78], 7, -8.5435, 'hello'],
            [
                ('qwe', 'bold_font_style'),
                ('[6, 7, 78]', 'DEFAULT_STYLE'),
                (7, 'DEFAULT_STYLE'),
                (-8.5435, 'DEFAULT_STYLE'),
                ('hello', 'DEFAULT_STYLE'),
            ],
        ),
        (-1, [1, 2, 3], ValueError),
        (1048576, [1], ValueError),
        (2, 99, ValueError),
        (6, 'STRING', ValueError),
    ],
)
def test_worsheet_row_get_and_set(index, value_list, expected_output):
    from pyfastexcel.utils import set_custom_style

    style = CustomStyle(font_size=12, font_bold=True)
    set_custom_style('bold_font_style', style)
    wb = Workbook()
    ws = wb['Sheet1']

    if not isinstance(expected_output, list):
        print('qweqweqw')
        with pytest.raises(expected_output):
            ws[index] = value_list
    else:
        ws[index] = value_list
        print(ws[index])
        assert ws[index] == expected_output


@pytest.mark.parametrize(
    'cell_value',
    [
        ([('1', '2', '3')]),
        (('1', 3, 2)),
        (('1')),
        ((1)),
    ],
)
def test_set_worksheet_with_wrong_format(cell_value):
    wb = Workbook()
    ws = wb['Sheet1']
    with pytest.raises(ValueError):
        ws[0] = cell_value


def test_save_workbook():
    wb = Workbook()
    ws = wb['Sheet1']
    ws['A1':'G1'] = [1, 2, 3, 9, 8, 45, 11]

    # Save without calling read_lib_and_create_excel()
    wb.save('test1.xlsx')

    wb.read_lib_and_create_excel()
    wb.save('test2.xlsx')


def test_if_style_is_reset():
    from pyfastexcel.driver import ExcelDriver

    wb = Workbook()
    style = CustomStyle(font_size=11, font_color='000000')

    ws = wb['Sheet1']
    ws['A1'] = ('test', style)
    wb._create_style()
    assert len(ExcelDriver._style_map) != 0
    assert len(ExcelDriver._STYLE_NAME_MAP) != 0
    assert ExcelDriver._STYLE_ID == 1
    assert ExcelDriver.REGISTERED_STYLES == {
        'DEFAULT_STYLE': ExcelDriver.DEFAULT_STYLE,
        'Custom Style 0': style,
    }
    wb.read_lib_and_create_excel()
    assert len(ExcelDriver._style_map) == 0
    assert len(ExcelDriver._STYLE_NAME_MAP) == 0
    assert ExcelDriver._STYLE_ID == 0
    assert ExcelDriver.REGISTERED_STYLES == {
        'DEFAULT_STYLE': ExcelDriver.DEFAULT_STYLE,
    }

    # Create another Workbook in one process to ensure that after style configs
    # reset, everythings is still working.
    wb2 = Workbook()
    style2 = CustomStyle(font_size=99, font_color='fcfcfc')

    ws2 = wb2['Sheet1']
    ws2['A1'] = ('test', style2)
    wb2._create_style()
    assert len(ExcelDriver._style_map) != 0
    assert len(ExcelDriver._STYLE_NAME_MAP) != 0
    assert ExcelDriver._STYLE_ID == 1
    assert ExcelDriver.REGISTERED_STYLES == {
        'DEFAULT_STYLE': ExcelDriver.DEFAULT_STYLE,
        'Custom Style 0': style2,
    }
    wb2.read_lib_and_create_excel()
    assert len(ExcelDriver._style_map) == 0
    assert len(ExcelDriver._STYLE_NAME_MAP) == 0
    assert ExcelDriver._STYLE_ID == 0
    assert ExcelDriver.REGISTERED_STYLES == {
        'DEFAULT_STYLE': ExcelDriver.DEFAULT_STYLE,
    }


@pytest.mark.parametrize(
    'target, expected_output1',
    [
        ('A1', ('test', 'bold_font_style')),
        ('XD1', ('test', 'bold_font_style')),
    ],
)
def test_set_style_with_str(target, expected_output1):
    from pyfastexcel.driver import ExcelDriver
    from pyfastexcel.utils import set_custom_style

    wb = Workbook()
    ws = wb['Sheet1']

    bold_style = CustomStyle(font_bold=True)
    set_custom_style('bold_font_style', bold_style)
    color_style = CustomStyle(font_color='d33513')

    ws[target] = 'test'

    ws.set_style(target, 'bold_font_style')
    assert ws[target] == expected_output1

    ws.set_style(target, color_style)
    assert ws[target][1] == f'Custom Style {ExcelDriver._STYLE_ID - 1}'

    with pytest.raises(ValueError):
        ws.set_style(target, 'wrong_style')


@pytest.mark.parametrize(
    'target, expected_output1',
    [
        ('A1:B1', [('test', 'bold_font_style'), ('q', 'bold_font_style')]),
        (slice('A1', 'B1'), [('test', 'bold_font_style'), ('q', 'bold_font_style')]),
    ],
)
def test_set_style_with_silce(target, expected_output1):
    from pyfastexcel.driver import ExcelDriver
    from pyfastexcel.utils import set_custom_style

    wb = Workbook()
    ws = wb['Sheet1']

    bold_style = CustomStyle(font_bold=True)
    set_custom_style('bold_font_style', bold_style)
    color_style = CustomStyle(font_color='d33513')

    # Index assignment is not supported for slice currently.
    if isinstance(target, str) and ':' in target:
        t = target.split(':')
        t = slice(t[0], t[1])
    else:
        t = target

    ws[t] = ['test', 'q']

    ws.set_style(target, 'bold_font_style')
    assert ws[t] == expected_output1

    ws.set_style(target, color_style)
    assert ws[t][1][1] == f'Custom Style {ExcelDriver._STYLE_ID - 1}'

    with pytest.raises(ValueError):
        ws.set_style(target, 'wrong_style')


@pytest.mark.parametrize(
    'row, target, expected_output1',
    [
        (0, [0, 1], [('test', 'DEFAULT_STYLE'), ('q', 'bold_font_style'), ('1', 'DEFAULT_STYLE')]),
        (1, [1, 1], [('test', 'DEFAULT_STYLE'), ('q', 'bold_font_style'), ('1', 'DEFAULT_STYLE')]),
    ],
)
def test_set_style_with_list(row, target, expected_output1):
    from pyfastexcel.driver import ExcelDriver
    from pyfastexcel.utils import set_custom_style

    wb = Workbook()
    ws = wb['Sheet1']

    bold_style = CustomStyle(font_bold=True)
    set_custom_style('bold_font_style', bold_style)
    color_style = CustomStyle(font_color='d33513')

    ws[row] = ['test', 'q', '1']

    ws.set_style(target, 'bold_font_style')
    assert ws[row] == expected_output1

    ws.set_style(target, color_style)
    assert ws[row][target[1]][1] == f'Custom Style {ExcelDriver._STYLE_ID - 1}'


@pytest.mark.parametrize(
    'target, expected_output',
    [
        ([1048577, 1], ValueError),
        ([1, 16385], ValueError),
        (['awer', 1], TypeError),
        ({}, TypeError),
    ],
)
def test_set_style_error(target, expected_output):
    from pyfastexcel.utils import set_custom_style

    wb = Workbook()
    ws = wb['Sheet1']

    bold_style = CustomStyle(font_bold=True)
    set_custom_style('bold_font_style', bold_style)

    with pytest.raises(expected_output):
        ws.set_style(target, 'bold_font_style')

    writer = NormalWriter({})
    with pytest.raises(IndexError):
        writer['Sheet1'].set_style(target, 'bold_font_style')
