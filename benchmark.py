import statistics
import timeit

from openpyxl import Workbook as OpenpyxlWorkbook

from example import prepare_example_data
from pyfastexcel import StreamWriter
from pyfastexcel import Workbook as PyFastExcelWorkbook

data = prepare_example_data(rows=5000, cols=30)


setup = """
from __main__ import write_excel_with_pyfastexcel_with_double_for_loop
from __main__ import write_excel_with_pyfastexcel_with_row
from __main__ import write_excel_with_stream_writer
from __main__ import write_excel_with_openpyxl_normal_wb
from __main__ import write_excel_with_openpyxl_write_only_wb
"""


def write_excel_with_pyfastexcel_with_double_for_loop() -> None:
    from pyfastexcel.utils import index_to_column

    wb = PyFastExcelWorkbook()
    ws = wb['Sheet1']
    ws[0] = list(data[0].keys())
    for i, record in enumerate(data):
        for j, (_, value) in enumerate(record.items()):
            col = index_to_column(j + 2)
            ws[f'{col}{i}'] = value
    wb.save('pyfastexcel_double_for_loop.xlsx')


def write_excel_with_pyfastexcel_with_row() -> None:
    wb = PyFastExcelWorkbook()
    ws = wb['Sheet1']
    for i, record in enumerate(data):
        ws[i] = list(record.values())
    wb.save('pyfastexcel_by_row.xlsx')


def write_excel_with_stream_writer() -> None:
    class CustomWriter(StreamWriter):
        def create_excel(self) -> bytes:
            self._set_header()
            self._create_style()
            self._create_body()
            self.save('pyfastexcel_stream_writer.xlsx')

        def _set_header(self):
            self.headers = list(self.data[0].keys())
            for h in self.headers:
                self.row_append(h)
            self.create_row()

        def _create_body(self) -> None:
            for row in self.data:
                for h in self.headers:
                    self.row_append(row[h])
                self.create_row()

    CustomWriter(data).create_excel()


def write_excel_with_openpyxl_normal_wb() -> None:
    wb = OpenpyxlWorkbook()
    ws = wb.create_sheet()
    ws.append(list(data[0].keys()))
    for record in data:
        ws.append(list(record.values()))
    wb.save('openpyxl_wb.xlsx')


def write_excel_with_openpyxl_write_only_wb() -> None:
    wb = OpenpyxlWorkbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(list(data[0].keys()))
    for record in data:
        ws.append(list(record.values()))
    wb.save('openpyxl_write_only_wb.xlsx')


def run_test_case(test_case_name, test_case_func, repeat=5, number=1):
    benchmark = f'\nExecution time ({test_case_name}):\n'
    results = timeit.repeat(
        f'{test_case_func}()',
        setup=setup,
        repeat=repeat,
        number=number,
    )
    for i, time in enumerate(results):
        benchmark += f'Test {i + 1}: {time} s\n'

    mean = statistics.mean(results)
    std_dev = statistics.stdev(results)
    max_val = max(results)
    min_val = min(results)
    benchmark += f'Mean: {mean} s\nMax: {max_val} s\nMin: {min_val} s\nStd: {std_dev} s\n'
    return benchmark


if __name__ == '__main__':
    benchmark = run_test_case(
        'WorkBook double loop',
        'write_excel_with_pyfastexcel_with_double_for_loop',
    )
    benchmark += run_test_case('WorkBook by row', 'write_excel_with_pyfastexcel_with_row')
    benchmark += run_test_case('StreamWriter', 'write_excel_with_stream_writer')
    benchmark += run_test_case('Openpyxl Workbook', 'write_excel_with_openpyxl_normal_wb')
    benchmark += run_test_case(
        'Openpyxl Write Only Workbook',
        'write_excel_with_openpyxl_write_only_wb',
    )
    print(benchmark)
