from __future__ import annotations

import time

from openpyxl.styles import Side

from pyfastexcel import CustomStyle, StreamWriter, Workbook
from pyfastexcel.utils import set_custom_style


def prepare_example_data(rows: int = 1000, cols: int = 10) -> list[dict[str, str]]:
    headers = [f'Column_{i}' for i in range(cols)]
    data = [[i for i in range(cols)] for j in range(rows)]
    records = []
    for row in data:
        record = {}
        for header, value in zip(headers, row):
            record[header] = str(round(value * 100, 2))
        records.append(record)
    return records


class StyleCollections:
    black_fill_style = CustomStyle(
        font_name='Time News Roman',
        font_size='11',
        font_bold=True,
        font_color='F62B00',
        fill_color='000000',
    )
    green_fill_style = CustomStyle(
        font_size='29',
        font_bold=False,
        font_color='000000',
        fill_color='375623',
    )
    test_fill_style = CustomStyle(
        font_params={
            'size': 20,
            'bold': True,
            'italic': True,
            'color': '5e03fc',
        },
        fill_params={
            'patternType': 'solid',
            'fgColor': '375623',
        },
        border_params={
            'left': Side(style='thin', color='e12aeb'),
            'right': Side(style='thick', color='e12aeb'),
            'top': Side(style=None, color='e12aeb'),
            'bottom': Side(style='dashDot', color='e12aeb'),
        },
        ali_params={
            'wrapText': True,
            'shrinkToFit': True,
        },
        number_format='0.00%',
    )


class PyFastExcelStreamExample(StreamWriter, StyleCollections):
    def create_excel(self) -> bytes:
        self._set_header()
        self._create_style()
        self.set_file_props('Creator', 'Hello')
        self._create_single_header()
        self._create_body()
        return self.read_lib_and_create_excel()

    def _set_header(self):
        self.headers = list(self.data[0].keys())

    def _create_single_header(self):
        for h in self.headers:
            self.row_append(h, style='green_fill_style')
        self.create_row()

    def _create_body(self) -> None:
        for row in self.data:
            for h in self.headers:
                if h[-1] in ('1', '3', '5', '7', '9'):
                    self.row_append(row[h], style='black_fill_style')
                else:
                    self.row_append(row[h], style='test_fill_style')
            self.create_row()

        self.create_sheet('Sheet2')
        for row in self.data:
            for h in self.headers:
                if h[-1] in ('1', '3', '5', '7', '9'):
                    self.row_append(row[h], style=self.green_fill_style)
                else:
                    self.row_append(row[h], style='black_fill_style')
            self.create_row()

        custom_style2 = CustomStyle(
            font_size='33',
            font_bold=True,
            font_color='000000',
            fill_color='4db3af',
        )
        self.row_append('New Style', style=custom_style2)
        self.create_row()
        self.set_cell_width(self.sheet, 'A', 255)
        self.set_cell_height(self.sheet, 4, 123)
        self.merge_cell(self.sheet, 'A2', 'A6')
        self.workbook['Sheet1']['A2'] = ('Hellow World', 'black_fill_style')
        self.workbook['Sheet1']['A3'] = 'I am A3'
        self.workbook['Sheet1']['AB9'] = 'qwer'


if __name__ == '__main__':
    data = prepare_example_data(6, 9)

    # StreamWriter
    normal_start_time = time.perf_counter()
    excel_normal = PyFastExcelStreamExample(data).create_excel()
    notmal_end_time = time.perf_counter()
    print('PyFastExcelStreamWriter time: ', notmal_end_time - normal_start_time)

    # Workbook
    wb = Workbook()

    # Set and register CustomStyle
    bold_style = CustomStyle(font_size=15, font_bold=True)
    set_custom_style('bold_style', bold_style)

    ws = wb['Sheet1']
    # Write value with default style
    ws['A1'] = 'A1 value'
    # Write value with custom style
    ws['B1'] = ('B1 value', 'bold_style')

    # Write value in slice with default style
    ws['A2':'C2'] = [1, 2, 3]
    # Write value in slice with custom style
    ws['A3':'C3'] = [(1, 'bold_style'), (2, 'bold_style'), (3, 'bold_style')]

    # Write value by row with default style (python index 0 is the index 1 in excel)
    ws[3] = [9, 8, 'go']
    # Write value by row with custom style
    ws[4] = [(9, 'bold_style'), (8, 'bold_style'), ('go', 'bold_style')]

    # Send request to golang lib and create excel
    wb.read_lib_and_create_excel()

    # File path to save
    file_path1 = 'pyexample_stream.xlsx'
    file_path2 = 'pyexample_workbook.xlsx'

    # Save to the xlsx with save function of Workbook
    wb.save(file_path2)

    # Save to the xlsx by the bytes return from create_excel()
    with open(file_path1, 'wb') as file:
        file.write(excel_normal)
