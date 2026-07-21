"""pyfastexcel vs openpyxl throughput benchmark.

Runs a size matrix of Workbook / StreamWriter / openpyxl writers and produces a
horizontal-bar chart per case.  Each run also writes a structured JSON result to
``benchmark/results/openpyxl/<date>-<os>-openpyxl.json`` so re-running never
overwrites earlier numbers; historical PNGs are likewise left untouched.

Usage:
    uv run python benchmark/benchmark.py                       # auto-detect OS
    uv run python benchmark/benchmark.py --os-name Windows11   # override label
    uv run python benchmark/benchmark.py --repeat 3
"""

from __future__ import annotations

import argparse
import datetime as dt
import importlib.metadata
import json
import os
import platform
import re
import statistics
import sys
import timeit
from pathlib import Path

import matplotlib

matplotlib.use('Agg')
import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402
from openpyxl import Workbook as OpenpyxlWorkbook  # noqa: E402

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))  # noqa

from example import prepare_example_data  # noqa: E402
from pyfastexcel import StreamWriter  # noqa: E402
from pyfastexcel import Workbook as PyFastExcelWorkbook  # noqa: E402

RESULTS_DIR = Path(__file__).resolve().parent / 'results' / 'openpyxl'
DEFAULT_CASES = [(50, 30), (500, 30), (5000, 30), (50000, 30)]

# Assigned in main(); the plotting helpers read them as module globals.
os_name = platform.system() or 'unknown'
os_title = f'OS: {os_name}'
data = None
result_dict = {}


def _detect_os_name() -> str:
    """A filesystem-friendly OS label, e.g. ``Windows11`` or ``Linux-6.18``."""
    system = platform.system()
    if system == 'Windows':
        return f'Windows{platform.release()}'
    if system == 'Linux':
        release = platform.release()
        wsl = 'WSL2-' if 'microsoft' in release.lower() else ''
        # Kept out of the f-string: nesting the same quote needs Python 3.12,
        # and the package supports 3.10.
        version = release.split('-')[0]
        return f'{wsl}Linux-{version}'
    if system == 'Darwin':
        return f'macOS-{platform.mac_ver()[0]}'
    return system or 'unknown'


def _package_version(distribution: str) -> str | None:
    try:
        return importlib.metadata.version(distribution)
    except importlib.metadata.PackageNotFoundError:
        return None


def _cpu_name() -> str | None:
    if processor := platform.processor():
        return processor
    try:
        for line in Path('/proc/cpuinfo').read_text(encoding='utf-8').splitlines():
            if line.lower().startswith('model name'):
                return line.partition(':')[2].strip()
    except OSError:
        pass
    return None


setup = """
from __main__ import write_excel_with_pyfastexcel_with_double_for_loop
from __main__ import write_excel_with_pyfastexcel_with_row
from __main__ import write_excel_with_stream_writer
from __main__ import write_excel_with_openpyxl_normal_wb
from __main__ import write_excel_with_openpyxl_write_only_wb
"""


def write_excel_with_pyfastexcel_with_double_for_loop() -> None:
    from pyfastexcel.utils import index_to_column

    wb = PyFastExcelWorkbook()
    ws = wb['Sheet1']
    ws[0] = list(data[0].keys())
    for i, record in enumerate(data):
        for j, (_, value) in enumerate(record.items()):
            col = index_to_column(j + 2)
            ws[f'{col}{i}'] = value
    wb.save('pyfastexcel_double_for_loop.xlsx')


def write_excel_with_pyfastexcel_with_row() -> None:
    wb = PyFastExcelWorkbook()
    ws = wb['Sheet1']
    for i, record in enumerate(data):
        ws[i] = list(record.values())
    wb.save('pyfastexcel_by_row.xlsx')


def write_excel_with_stream_writer() -> None:
    class CustomWriter(StreamWriter):
        def create_excel(self) -> bytes:
            self._set_header()
            self._create_style()
            self._create_body()
            self.save('pyfastexcel_stream_writer.xlsx')

        def _set_header(self):
            self.headers = list(self.data[0].keys())
            for h in self.headers:
                self.row_append(h)
            self.create_row()

        def _create_body(self) -> None:
            for row in self.data:
                for h in self.headers:
                    self.row_append(row[h])
                self.create_row()

    CustomWriter(data).create_excel()


def write_excel_with_openpyxl_normal_wb() -> None:
    wb = OpenpyxlWorkbook()
    ws = wb.create_sheet()
    ws.append(list(data[0].keys()))
    for record in data:
        ws.append(list(record.values()))
    wb.save('openpyxl_wb.xlsx')


def write_excel_with_openpyxl_write_only_wb() -> None:
    wb = OpenpyxlWorkbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(list(data[0].keys()))
    for record in data:
        ws.append(list(record.values()))
    wb.save('openpyxl_write_only_wb.xlsx')


def run_test_case(test_case_name, test_case_func, repeat=5, number=1):
    benchmark = f'\nExecution time ({test_case_name}): \n'
    results = timeit.repeat(
        f'{test_case_func}()',
        setup=setup,
        repeat=repeat,
        number=number,
    )
    result_dict[test_case_name] = {'results': results}
    for i, time in enumerate(results):
        benchmark += f'Test {i + 1}: {time} s\n'

    mean = statistics.mean(results)
    std_dev = statistics.stdev(results)
    max_val = max(results)
    min_val = min(results)
    benchmark += f'Mean: {mean} s\nMax: {max_val} s\nMin: {min_val} s\nStd: {std_dev} s\n'
    result_dict[test_case_name]['mean'] = mean
    result_dict[test_case_name]['max_val'] = max_val
    result_dict[test_case_name]['min_val'] = min_val
    result_dict[test_case_name]['std_dev'] = std_dev
    return benchmark


def extract_plot_data(result_dict):
    labels = list(result_dict.keys())
    means = [entry['mean'] for entry in result_dict.values()]
    max_vals = [entry['max_val'] for entry in result_dict.values()]
    min_vals = [entry['min_val'] for entry in result_dict.values()]
    std_devs = [entry['std_dev'] for entry in result_dict.values()]
    return labels, means, max_vals, min_vals, std_devs


def add_labels(bars, ax, orientation='v'):
    for bar in bars:
        if orientation == 'v':
            height = bar.get_height()
            ax.annotate(
                f'{height:.3f}',
                xy=(bar.get_x() + bar.get_width() / 2, height),
                xytext=(0, 3),  # 3 points vertical offset
                textcoords='offset points',
                ha='center',
                va='bottom',
            )
        else:
            width = bar.get_width()
            ax.annotate(
                f'{width:.3f}',
                xy=(width, bar.get_y() + bar.get_height() / 2),
                xytext=(3, 0),  # 3 points horizontal offset
                textcoords='offset points',
                ha='left',
                va='center',
            )


def plot_bars(orientation='v', title='Method', fig_name='benchmark.png'):
    labels, means, max_vals, min_vals, std_devs = extract_plot_data(result_dict)
    num_bars = len(labels)
    indices = np.arange(num_bars)
    bar_width = 0.2

    _, ax = plt.subplots(figsize=(10, 6))

    if orientation == 'v':
        bars1 = ax.bar(
            indices - bar_width,
            min_vals,
            bar_width,
            label='Min',
            color='orange',
            edgecolor='black',
        )
        bars2 = ax.bar(indices, means, bar_width, label='Mean', color='cyan', edgecolor='black')
        bars3 = ax.bar(
            indices + bar_width,
            max_vals,
            bar_width,
            label='Max',
            color='gray',
            edgecolor='black',
        )
        ax2 = ax.twinx()
        ax2.plot(indices, std_devs, '--o', color='red', label='Std Dev')
        ax.set_xlabel(title, fontsize=14)
        ax.set_ylabel('Time (s)', fontsize=14)
        ax.set_xticks(indices)
        ax.set_xticklabels(labels, fontsize=10)
        ax2.set_ylabel('Standard Deviation')
    else:
        bars1 = ax.barh(
            indices - bar_width,
            min_vals,
            bar_width,
            label='Min',
            color='orange',
            edgecolor='black',
        )
        bars2 = ax.barh(indices, means, bar_width, label='Mean', color='cyan', edgecolor='black')
        bars3 = ax.barh(
            indices + bar_width,
            max_vals,
            bar_width,
            label='Max',
            color='gray',
            edgecolor='black',
        )
        bars4 = ax.barh(
            indices + 2 * bar_width,
            std_devs,
            bar_width,
            label='Std',
            color='pink',
            edgecolor='black',
        )
        ax.set_ylabel(title, fontsize=14)
        ax.set_xlabel('Time (s)', fontsize=14)
        ax.set_yticks(indices)
        ax.set_yticklabels(labels)

    ax.set_title(os_title, fontsize=14)
    handles1, labels1 = ax.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels() if orientation == 'v' else ([], [])
    ax.legend(handles1 + handles2, labels1 + labels2)

    add_labels(bars1, ax, orientation)
    add_labels(bars2, ax, orientation)
    add_labels(bars3, ax, orientation)
    if orientation == 'h':
        add_labels(bars4, ax, orientation)

    if orientation == 'v':
        for i, std in enumerate(std_devs):
            ax2.annotate(
                f'{std:.3f}',
                xy=(indices[i], std),
                xytext=(0, 5),
                textcoords='offset points',
                ha='center',
                va='bottom',
            )

    plt.tight_layout()
    plt.savefig(fig_name)


def plot_vertical_bar(title='Method', fig_name='vbars.png'):
    plot_bars(orientation='v', title=title, fig_name=fig_name)


def plot_horizontal_bar(title='Method', fig_name='hbars.png'):
    plot_bars(orientation='h', title=title, fig_name=fig_name)


def _parse_cases(spec: str | None) -> list[tuple[int, int]]:
    if not spec:
        return DEFAULT_CASES
    cases = []
    for pair in spec.split(','):
        match = re.fullmatch(r'\s*(\d+)x(\d+)\s*', pair)
        if not match:
            raise SystemExit(f'invalid --cases entry {pair!r}; expected e.g. "500x30,5000x30"')
        cases.append((int(match.group(1)), int(match.group(2))))
    return cases


def _save_results_json(all_cases: list[dict], repeat: int, output_dir: Path) -> Path:
    today = dt.date.today().isoformat()
    report = {
        'date': today,
        'os_name': os_name,
        'repeat': repeat,
        'environment': {
            'platform': platform.platform(),
            'cpu': _cpu_name(),
            'python': platform.python_version(),
            'openpyxl': _package_version('openpyxl'),
            'pyfastexcel': _package_version('pyfastexcel'),
        },
        'cases': all_cases,
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f'{today}-{os_name}-openpyxl.json'
    path.write_text(json.dumps(report, indent=2) + '\n', encoding='utf-8')
    return path


def main() -> None:
    global data, os_name, os_title, result_dict

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        '--os-name', default=_detect_os_name(), help='OS label for titles/filenames'
    )
    parser.add_argument('--repeat', type=int, default=5, help='timeit repeats per case')
    parser.add_argument('--cases', help='comma-separated ROWSxCOLS list, e.g. "500x30,5000x30"')
    parser.add_argument(
        '--output-dir',
        type=Path,
        default=Path(__file__).resolve().parent,
        help='directory for the per-case PNGs (default: benchmark/)',
    )
    args = parser.parse_args()

    os_name = args.os_name
    os_title = f'OS: {os_name}'
    cases = _parse_cases(args.cases)

    methods = [
        ('WorkBook double loop', 'write_excel_with_pyfastexcel_with_double_for_loop'),
        ('WorkBook by row', 'write_excel_with_pyfastexcel_with_row'),
        ('StreamWriter', 'write_excel_with_stream_writer'),
        ('Openpyxl\nWorkbook', 'write_excel_with_openpyxl_normal_wb'),
        ('Openpyxl Write\nOnly Workbook', 'write_excel_with_openpyxl_write_only_wb'),
    ]

    all_cases = []
    for row, col in cases:
        data = prepare_example_data(rows=row, cols=col)
        result_dict = {}
        benchmark = ''
        for label, func in methods:
            benchmark += run_test_case(label, func, repeat=args.repeat)
        print(benchmark)
        plot_horizontal_bar(
            f'Method (rows={row}, columns={col})',
            str(args.output_dir / f'{row}_{col}_horizontal_{os_name}.png'),
        )
        all_cases.append({'rows': row, 'cols': col, 'results': result_dict})

    saved = _save_results_json(all_cases, args.repeat, RESULTS_DIR)
    print(f'\nSaved results JSON to {saved}')


if __name__ == '__main__':
    main()
